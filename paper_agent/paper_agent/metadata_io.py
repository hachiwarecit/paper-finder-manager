"""人間が編集した Excel / CSV からメタデータを取り込み、PaperRecord を更新する。

ワークフロー:
    1. export --format xlsx で paper_inventory.xlsx を出力
    2. 人間が Excel を開いて authors / sample_size / doi 等を補完
    3. import-metadata で編集内容を SQLite に反映
    4. dedupe-all / screen-all / report で再判定

方針:
    - 更新対象は EDITABLE_COLUMNS のみ。screening_status などの判定結果列は触らない。
    - **空欄（空文字 / NaN / None）は既存値を上書きしない**（補完であって消去ではない）。
    - paper_id をキーに既存レコードを探す。存在しない paper_id はスキップ（警告）。
    - 不正な値（数値列に文字、document_type に未知の値）はスキップして警告。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from .db import PaperDB
from .models import CandidateStatus, DocumentType, PaperRecord
from .utils import get_logger, normalize_title

logger = get_logger()

# Candidates シートで人間が編集して取り込める列 -> Candidate フィールド
CANDIDATE_EDITABLE_COLUMNS: dict[str, str] = {
    "candidate_status": "candidate_status",
    "notes": "notes",
    "category": "category",
    "target_country": "target_country",
    "generation_keywords": "generation_keywords",
    "workplace_keywords": "workplace_keywords",
    "document_type_guess": "document_type_guess",
    "open_access_flag": "open_access_flag",
    "pdf_url": "pdf_url",
}
_BOOL_TRUE = {"yes", "true", "1", "y", "t"}
_BOOL_FALSE = {"no", "false", "0", "n", "f", ""}

# 取り込み対象列 -> PaperRecord フィールド名
# Excel 列名 "method" は research_method に対応（export 側の列名にあわせる）。
EDITABLE_COLUMNS: dict[str, str] = {
    "title": "title",
    "authors": "authors",
    "year": "year",
    "doi": "doi",
    "country": "country",
    "category": "category",
    "target_country": "target_country",
    "organization_context": "organization_context",
    "generation_groups": "generation_groups",
    "number_of_generations": "number_of_generations",
    "sample_size": "sample_size",
    "method": "research_method",
    "research_method": "research_method",
    "document_type": "document_type",
    "original_language": "original_language",
    "analysis_language": "analysis_language",
    "notes": "notes",
}

_INT_FIELDS = {"year", "number_of_generations", "sample_size"}


@dataclass
class ImportStats:
    rows_read: int = 0
    records_updated: int = 0
    fields_updated: int = 0
    candidates_updated: int = 0
    candidate_fields_updated: int = 0
    skipped_missing_id: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _is_blank(value) -> bool:
    """空欄判定: None / 空文字 / 空白のみ / NaN を空とみなす。"""
    if value is None:
        return True
    # pandas/float NaN (NaN != NaN)
    if isinstance(value, float) and value != value:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _read_xlsx_sheets(p: Path) -> dict[str, list[dict]]:
    """xlsx の全シートを {sheet_name: rows} で返す。"""
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
    result: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            result[name] = []
            continue
        header = [str(h).strip() if h is not None else "" for h in header]
        rows = []
        for row in rows_iter:
            rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
        result[name] = rows
    wb.close()
    return result


def _read_csv(p: Path) -> list[dict]:
    import csv

    # Excel が書く UTF-8 BOM (utf-8-sig) にも対応
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(r) for r in reader]


def _coerce(field_name: str, value, stats: ImportStats, paper_id: str):
    """フィールド型に合わせて値を変換。失敗時は (None, False)。成功時 (val, True)。"""
    if field_name in _INT_FIELDS:
        try:
            return int(float(str(value).strip())), True
        except (TypeError, ValueError):
            stats.warnings.append(f"{paper_id}: {field_name} の値 '{value}' は数値でないため無視。")
            return None, False
    if field_name == "document_type":
        v = str(value).strip()
        valid = {d.value for d in DocumentType}
        if v not in valid:
            stats.warnings.append(
                f"{paper_id}: document_type '{v}' は不正。許可値: {sorted(valid)}。無視。"
            )
            return None, False
        return DocumentType(v), True
    return str(value).strip(), True


def _apply_paper_rows(rows: list[dict], db: PaperDB, stats: ImportStats) -> None:
    for row in rows:
        pid = row.get("paper_id")
        if _is_blank(pid):
            continue
        pid = str(pid).strip()
        rec = db.get(pid)
        if rec is None:
            stats.skipped_missing_id.append(pid)
            stats.warnings.append(f"{pid}: DB に存在しない paper_id のためスキップ。")
            continue

        changed = False
        title_changed = False
        for col, value in row.items():
            if col not in EDITABLE_COLUMNS:
                continue
            if _is_blank(value):
                continue  # 空欄は既存値を保持
            field_name = EDITABLE_COLUMNS[col]
            coerced, ok = _coerce(field_name, value, stats, pid)
            if not ok:
                continue
            current = getattr(rec, field_name)
            cur_cmp = current.value if hasattr(current, "value") else current
            new_cmp = coerced.value if hasattr(coerced, "value") else coerced
            if str(cur_cmp) == str(new_cmp):
                continue
            setattr(rec, field_name, coerced)
            stats.fields_updated += 1
            changed = True
            if field_name == "title":
                title_changed = True

        if title_changed:
            rec.normalized_title = normalize_title(rec.title)
        if changed:
            db.upsert(rec)
            stats.records_updated += 1


def _coerce_candidate(field_name: str, value, stats: ImportStats, cid: str):
    """Candidate フィールド向けの値変換。失敗時 (None, False)。"""
    if field_name == "candidate_status":
        v = str(value).strip()
        valid = {s.value for s in CandidateStatus}
        if v not in valid:
            stats.warnings.append(
                f"{cid}: candidate_status '{v}' は不正。許可値: {sorted(valid)}。無視。"
            )
            return None, False
        return CandidateStatus(v), True
    if field_name == "open_access_flag":
        v = str(value).strip().lower()
        if v in _BOOL_TRUE:
            return True, True
        if v in _BOOL_FALSE:
            return False, True
        stats.warnings.append(f"{cid}: open_access_flag '{value}' は真偽値でないため無視。")
        return None, False
    return str(value).strip(), True


def _apply_candidate_rows(rows: list[dict], db: PaperDB, stats: ImportStats) -> None:
    for row in rows:
        cid = row.get("candidate_id")
        if _is_blank(cid):
            continue
        cid = str(cid).strip()
        cand = db.get_candidate(cid)
        if cand is None:
            stats.skipped_missing_id.append(cid)
            stats.warnings.append(f"{cid}: DB に存在しない candidate_id のためスキップ。")
            continue

        changed = False
        for col, value in row.items():
            if col not in CANDIDATE_EDITABLE_COLUMNS:
                continue
            if _is_blank(value):
                continue
            field_name = CANDIDATE_EDITABLE_COLUMNS[col]
            coerced, ok = _coerce_candidate(field_name, value, stats, cid)
            if not ok:
                continue
            current = getattr(cand, field_name)
            cur_cmp = current.value if hasattr(current, "value") else current
            new_cmp = coerced.value if hasattr(coerced, "value") else coerced
            if str(cur_cmp) == str(new_cmp):
                continue
            setattr(cand, field_name, coerced)
            stats.candidate_fields_updated += 1
            changed = True
        if changed:
            db.upsert_candidate(cand)
            stats.candidates_updated += 1


def import_metadata(path: str | Path, db: PaperDB | None = None) -> ImportStats:
    """編集済み Excel/CSV を読み込み、空欄以外の編集列で既存レコードを更新する。

    Excel は All_Papers シート (論文) と Candidates シート (候補) の両方を取り込む。
    CSV はヘッダに paper_id があれば論文、candidate_id があれば候補として扱う。
    """
    stats = ImportStats()
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {p}")

    suffix = p.suffix.lower()
    paper_rows: list[dict] = []
    candidate_rows: list[dict] = []

    if suffix in (".xlsx", ".xlsm"):
        sheets = _read_xlsx_sheets(p)
        for name, rows in sheets.items():
            if not rows:
                continue
            cols = set(rows[0].keys())
            if "candidate_id" in cols:
                candidate_rows += rows
            elif "paper_id" in cols:
                paper_rows += rows
    elif suffix in (".csv", ".txt"):
        rows = _read_csv(p)
        if rows and "candidate_id" in rows[0]:
            candidate_rows = rows
        else:
            paper_rows = rows
    else:
        raise ValueError(f"対応していない形式です: {suffix} (xlsx / csv)")

    stats.rows_read = len(paper_rows) + len(candidate_rows)

    own = db is None
    db = db or PaperDB()
    try:
        _apply_paper_rows(paper_rows, db, stats)
        _apply_candidate_rows(candidate_rows, db, stats)
    finally:
        if own:
            db.close()

    logger.info(
        "import-metadata: %d 行読込 | 論文 %d 更新(%d列) | 候補 %d 更新(%d列)",
        stats.rows_read, stats.records_updated, stats.fields_updated,
        stats.candidates_updated, stats.candidate_fields_updated,
    )
    return stats
